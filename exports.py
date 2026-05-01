"""Excel and PDF export for event-based amortization schedules."""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)

from amortization import ScheduleRow


HEADERS = [
    "#", "Date", "Type", "Description",
    "Cash Flow", "Interest", "Principal", "Balance", "Accrued Int.",
]


def _fmt_money(x: float) -> str:
    return f"${x:,.2f}"


def to_excel(
    rows: list[ScheduleRow],
    summary: dict,
    loan_meta: dict,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amortization"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")

    ws["A1"] = loan_meta.get("title", "Loan Amortization Schedule")
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")

    info_pairs = [
        ("Borrower / Label", loan_meta.get("borrower", "")),
        ("Nominal Annual Rate", f"{loan_meta.get('nominal_annual_rate', 0):.4f}%"),
        ("Day Count", loan_meta.get("day_count", "")),
        ("Compounding", loan_meta.get("compounding", "")),
        ("First Event", str(loan_meta.get("first_date", "") or "")),
        ("Last Event", str(loan_meta.get("last_date", "") or "")),
    ]
    row_i = 3
    for lab, val in info_pairs:
        ws.cell(row=row_i, column=1, value=lab).font = label_font
        ws.cell(row=row_i, column=2, value=val)
        row_i += 1

    row_i += 1
    totals = [
        ("Total Disbursed", _fmt_money(summary["total_disbursed"])),
        ("Total Paid", _fmt_money(summary["total_paid"])),
        ("Total Interest", _fmt_money(summary["total_interest_paid"])),
        ("Total Principal Repaid", _fmt_money(summary["total_principal_paid"])),
        ("Ending Balance", _fmt_money(summary["ending_balance"])),
        ("Ending Accrued Interest", _fmt_money(summary["ending_accrued_interest"])),
        ("Net Cost (Paid − Disbursed)", _fmt_money(summary["net_cost"])),
    ]
    for lab, val in totals:
        ws.cell(row=row_i, column=1, value=lab).font = label_font
        ws.cell(row=row_i, column=2, value=val)
        row_i += 1

    # schedule header
    row_i += 2
    header_row = row_i
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=row_i, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    row_i += 1

    for r in rows:
        values = [
            r.seq, r.date, r.kind, r.description,
            r.cash_flow, r.interest, r.principal, r.balance, r.accrued_interest,
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.border = border
            if col == 1:
                c.alignment = center
            elif col == 2:
                c.number_format = "mm/dd/yyyy"
                c.alignment = center
            elif col in (3, 4):
                c.alignment = Alignment(horizontal="left")
            else:
                c.number_format = '"$"#,##0.00'
                c.alignment = right
        row_i += 1

    # totals row
    total_disbursed = sum(r.cash_flow for r in rows if r.kind == "Loan")
    total_paid = sum(r.cash_flow for r in rows if r.kind == "Payment")
    total_interest = sum(r.interest for r in rows if r.kind == "Payment")
    total_principal = sum(r.principal for r in rows if r.kind == "Payment")
    ws.cell(row=row_i, column=1, value="Totals").font = label_font
    for col, val in [
        (5, None),  # cash flow not summed (mixes loan & payment)
        (6, total_interest),
        (7, total_principal),
    ]:
        if val is None:
            continue
        cc = ws.cell(row=row_i, column=col, value=val)
        cc.number_format = '"$"#,##0.00'
        cc.font = label_font
        cc.alignment = right
    ws.cell(row=row_i, column=4,
            value=f"Disbursed {_fmt_money(total_disbursed)}, Paid {_fmt_money(total_paid)}").font = label_font

    widths = [6, 13, 10, 38, 14, 13, 13, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(
    rows: list[ScheduleRow],
    summary: dict,
    loan_meta: dict,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        leftMargin=0.4 * inch, rightMargin=0.4 * inch,
        topMargin=0.4 * inch, bottomMargin=0.4 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, spaceAfter=6)

    elements = []
    elements.append(Paragraph(loan_meta.get("title", "Loan Amortization Schedule"), title_style))

    info = [
        ["Borrower:", loan_meta.get("borrower", ""),
         "Rate:", f"{loan_meta.get('nominal_annual_rate', 0):.4f}%",
         "Day Count:", loan_meta.get("day_count", "")],
        ["Compounding:", loan_meta.get("compounding", ""),
         "First Event:", str(loan_meta.get("first_date", "") or ""),
         "Last Event:", str(loan_meta.get("last_date", "") or "")],
    ]
    info_tbl = Table(info, colWidths=[0.9*inch, 1.6*inch, 0.9*inch, 1.6*inch, 1.0*inch, 1.6*inch])
    info_tbl.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 9),
        ("FONT", (2, 0), (2, -1), "Helvetica-Bold", 9),
        ("FONT", (4, 0), (4, -1), "Helvetica-Bold", 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 8))

    tot_rows = [
        ["Total Disbursed:", _fmt_money(summary["total_disbursed"]),
         "Total Paid:", _fmt_money(summary["total_paid"]),
         "Total Interest:", _fmt_money(summary["total_interest_paid"])],
        ["Ending Balance:", _fmt_money(summary["ending_balance"]),
         "Accrued Int.:", _fmt_money(summary["ending_accrued_interest"]),
         "Net Cost:", _fmt_money(summary["net_cost"])],
    ]
    tot_tbl = Table(tot_rows, colWidths=[1.1*inch, 1.6*inch, 1.1*inch, 1.4*inch, 1.3*inch, 1.6*inch])
    tot_tbl.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 9),
        ("FONT", (2, 0), (2, -1), "Helvetica-Bold", 9),
        ("FONT", (4, 0), (4, -1), "Helvetica-Bold", 9),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F2F2")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tot_tbl)
    elements.append(Spacer(1, 10))

    data = [HEADERS]
    for r in rows:
        data.append([
            str(r.seq),
            r.date.strftime("%m/%d/%Y"),
            r.kind,
            r.description,
            _fmt_money(r.cash_flow),
            _fmt_money(r.interest),
            _fmt_money(r.principal),
            _fmt_money(r.balance),
            _fmt_money(r.accrued_interest),
        ])

    col_widths = [0.35*inch, 0.85*inch, 0.6*inch, 2.5*inch,
                  1.0*inch, 0.95*inch, 0.95*inch, 1.05*inch, 1.05*inch]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 8),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 7.5),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 0), (2, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "LEFT"),
        ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
    ]))
    elements.append(tbl)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(0.4 * inch, 0.25 * inch, f"Generated {date.today().isoformat()}")
        canvas.drawRightString(doc.pagesize[0] - 0.4 * inch, 0.25 * inch, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    return buf.getvalue()
